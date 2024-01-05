import html
from io import BytesIO
import re
import json
import datetime
from copy import copy
import openpyxl
from jsonschema import validate
from google.cloud import storage
from jsonschema import Draft202012Validator
import smtplib
from jinja2 import Template, Environment, FileSystemLoader
from email.message import EmailMessage

def save_obj_to_file(obj, save_path):
    with open(save_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=4, ensure_ascii=False)

def filter(node):
    if isinstance(node, dict):
        retVal = {}
        for key in node:
            if key == 'desc':
                if '#text' in node[key][0]:
                    retVal['description'] = html.unescape(node[key][0]['#text'])
            elif key == 'conformance':
                retVal['mro'] = node[key]
            elif key == 'shortName':
                retVal['name'] = node[key]
            elif key == 'operationalization':
                retVal['valueSets'] = html.unescape(node[key][0]['#text'])
            elif key == 'minimumMultiplicity':
                retVal[key] = node[key]
            elif key == 'maximumMultiplicity':
                retVal[key] = node[key]
            elif key == 'type': 
                retVal['type'] = node[key]
            elif key == 'valueDomain':
                if node[key][0]['type'] != 'code':
                    if node[key][0]['type'] != 'ordinal':
                        retVal[key] = copy(node[key])
            elif key == 'context':
                retVal['implementationGuidance'] = html.unescape(node[key][0]['#text'])
            elif isinstance(node[key], dict) or isinstance(node[key], list):
                if key not in ['relationship', 'implementation']:
                    child = filter(node[key])
                    if child:
                        retVal[key] = child
            
        if retVal:
            return retVal
        else:
            return None


    elif isinstance(node, list):
        retVal = []
        for entry in node:
            if isinstance(entry, str):
                retVal.append(entry)
            elif isinstance(entry, dict) or isinstance(entry, list):
                child = filter(entry)
                if child:
                    retVal.append(child)
        if retVal:
            return retVal
        else:
            return None



def validate(schema: dict, instance: dict):
    '''
    A function to validate the instance against a schema

    :param schema: the schema to validate against
    :param instance: the instance to validate
    '''
    validator = Draft202012Validator(schema)
    error_list = list(validator.iter_errors(instance))
    if error_list:
        return error_list
    else:
        return 'The instance is valid'

def load_worksheet(path, sheet_name, data_only, ret_type):
    '''
    A function that loads a worksheet from a workbook and returns an iterable

    :param path: the path to the file
    :param sheet_name: name of the sheet to load
    :param data_only: boolian value. Weather to load the sheet with cell formulae or values
    :param ret_type: weather it should be a python list or an openpyxl iterable. (list, excel)

    :return list: The  sheet as an iterable
    '''
    # open work book object
    wb = openpyxl.load_workbook(path, data_only=data_only)

    if ret_type == 'list':
        # load work sheet
        ws = list(wb[sheet_name])

        # translate into python dict
        full_data = []
        for arr in ws:
            data_arr = []
            for data in arr:
                data_arr.append(data.value)
            full_data.append(data_arr)
        return full_data

    else:
        # return openpyxl worksheet object
        ws = wb[sheet_name]
        return ws

def validate_contents(node, ws):
    if isinstance(node, dict):
        for key in node:
            if not isinstance(node[key], (dict, list)):
                if isinstance(node[key], str):
                    raw = node[key].strip().replace(' ','').lower().replace('_','')
                    found = False
                    for arr in ws:
                        for item in arr:
                            if str(item) != 'None':
                                raw_item = item.strip().replace(' ','').lower()
                                if raw == raw_item:
                                    print(f'Found {raw}')
                                    found = True
                                    break
                        if found:
                            break
                        
                    if not found:
                        print(f'Not found {raw}')
                else:
                    found = False
                    for arr in ws:
                        for item in arr:
                            if str(item) != 'None':
                                raw_item = item.strip().replace(' ','').lower()
                                if node[key] == raw_item:
                                    print(f'Found {node[key]}')
                                    found = True
                                    break
                        if found:
                            break
                    if not found:
                        print(f'Not found {node[key]}')
            else:
                validate_contents(node[key], ws)
    elif isinstance(node, list):
        for item in node:
            validate_contents(item, ws)

def send_email(subject, content, to_email, from_email, password):

    # Create the email message
    msg = EmailMessage()
    msg.set_content(content, subtype='html')
    msg['Subject'] = subject
    msg['From'] = from_email
    msg['To'] = to_email

    # Connect to the SMTP server
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(from_email, password)
        smtp.send_message(msg)


# a function to render a jinja template
def render_email(template_path, template_name, content, total_link_count, invalid_link_count, valid_link_count):
    # create jinja environment
    env = Environment(loader=FileSystemLoader(template_path))

    # load template
    template = env.get_template(template_name)

    # render template
    return template.render(standards=content, total_link_count=total_link_count, invalid_link_count=invalid_link_count, valid_link_count=valid_link_count, date=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))


def get_excel_from_gcloud(bucket_name: str, excel_name: str) -> openpyxl.Workbook:
    '''
    A function that will download the excel data from Google cloud without storing it locally
    
    :param bucket_name: the name of the bucket to download from
    :param excel_name: the name of the excel file to download
    :return: the excel file as an openpyxl workbook object
    '''
    client = storage.Client()
    bucket = client.get_bucket(bucket_name)
    blob = bucket.blob(excel_name)
    byte_data = BytesIO()
    blob.download_to_file(byte_data)
    byte_data.seek(0)
    wb = openpyxl.load_workbook(filename=byte_data, data_only=True)
    
    return wb

def find_cell_location(sheet, cell_value: str) -> tuple:
    '''
    Find the coordinates of a cell in a worksheet and return as a tuple.

    :param sheet: The sheet to search.
    :param cell_value: The text to search for in the sheet.
    :return: A tuple containing the row and column coordinates.
    '''
    row_index = 1
    for row in sheet.iter_rows():
        cell_index = 1
        for cell in row:
            value = str(cell.value).lower().strip()
            if value == cell_value.lower():
                return row_index, cell_index
            cell_index += 1
        row_index += 1
    raise ValueError(f"Missing headings in excel file {sheet.sheet_name}")

def get_endpoints_from_excel(workbook: openpyxl.Workbook) -> list:
    '''
    A function that will get the endpoints from an excel file
    
    :param workbook: the workbook to get the endpoints from
    :return: a list of endpoints
    '''
    ws = workbook['Sheet1']
    
    endpoint_heading_cell = find_cell_location(ws, 'Endpoints')
    
    endpoints = []
    for row in ws.iter_rows(min_row=endpoint_heading_cell[0]+1, min_col=endpoint_heading_cell[1], max_col=endpoint_heading_cell[1]):
        for cell in row:
            if cell.value:
                endpoints.append(cell.value)
    
    return endpoints
